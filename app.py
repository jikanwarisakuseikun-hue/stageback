import streamlit as st
from PIL import Image
import numpy as np
import os
import shutil
import pandas as pd
import zipfile
import colorsys
import openpyxl
from openpyxl.styles import Alignment, Border, Side, Font

# --- ページ基本設定 ---
st.set_page_config(
    page_title="カスタマイズ可能モザイクアートクリエーター",
    layout="wide"
)

# --- 用紙サイズの定義 (単位: mm, 余白を除いた目安の印刷可能エリア) ---
# ※excel_code には Excel標準の用紙サイズID（文字列）を指定してエラーを回避
PAPER_SIZES_MM = {
    "A4": {"width": 180.0, "height": 260.0, "excel_code": "9"},   # 9 = A4
    "A3": {"width": 260.0, "height": 380.0, "excel_code": "8"},   # 8 = A3
    "B5": {"width": 150.0, "height": 220.0, "excel_code": "13"},  # 13 = B5
    "B4": {"width": 220.0, "height": 330.0, "excel_code": "12"},  # 12 = B4
}

# --- 便利関数 ---

def hex_to_rgb(hex_color):
    """HEXコード(#RRGGBB)をRGBタプル(0-255)に変換"""
    hex_color = hex_color.lstrip('#')
    return tuple(int(hex_color[i:i+2], 16) for i in (0, 2, 4))

def rgb_to_hsv_tuple(rgb):
    """RGBタプル(0-255)をHSVタプル(0.0-1.0)に変換"""
    r, g, b = [x / 255.0 for x in rgb]
    return colorsys.rgb_to_hsv(r, g, b)

def create_mosaic_sheets_from_colors(
    base_image,
    target_colors_data, # [(HEX, 色名), ...]
    total_sheets_vertical,
    total_sheets_horizontal,
    tiles_per_sheet_vertical,
    tiles_per_sheet_horizontal,
    paper_size_name="A4",
    paper_orientation="タテ"
):
    tile_px_size = 10 # 仮想ピクセルサイズ

    # 全体のピクセル解像度計算
    mosaic_width_px = total_sheets_horizontal * tiles_per_sheet_horizontal * tile_px_size
    mosaic_height_px = total_sheets_vertical * tiles_per_sheet_vertical * tile_px_size

    base_img = base_image.convert("RGB")
    base_img_resized = base_img.resize((mosaic_width_px, mosaic_height_px), Image.Resampling.LANCZOS)

    target_colors_rgb = [hex_to_rgb(data[0]) for data in target_colors_data]
    target_color_names = [data[1] for data in target_colors_data]

    if not target_colors_rgb:
        raise ValueError("使用する色を少なくとも1つ指定してください。")

    # HSV空間での距離計算用データ作成（変な色の混ざりを防止）
    target_colors_hsv = np.array([rgb_to_hsv_tuple(rgb) for rgb in target_colors_rgb])
    
    output_mosaic_img = Image.new("RGB", (mosaic_width_px, mosaic_height_px))

    mosaic_color_names_grid = [
        ['' for _ in range(mosaic_width_px // tile_px_size)]
        for _ in range(mosaic_height_px // tile_px_size)
    ]

    # モザイク画像およびグリッドデータ生成
    tile_row_idx = 0
    for y in range(0, mosaic_height_px, tile_px_size):
        tile_col_idx = 0
        for x in range(0, mosaic_width_px, tile_px_size):
            tile_region = base_img_resized.crop((x, y, x + tile_px_size, y + tile_px_size))
            if tile_region.size[0] == 0 or tile_region.size[1] == 0:
                tile_col_idx += 1
                continue

            # タイル平均色をHSVに変換して比較
            tile_avg_rgb = np.array(tile_region).mean(axis=(0, 1))
            tile_avg_hsv = np.array(rgb_to_hsv_tuple(tile_avg_rgb))

            # HSV空間で距離が最も近い色を選択
            distances = np.linalg.norm(target_colors_hsv - tile_avg_hsv, axis=1)
            best_color_index = np.argmin(distances)

            # 画像に貼り付け
            selected_color_rgb = tuple(target_colors_rgb[best_color_index])
            color_tile = Image.new("RGB", (tile_px_size, tile_px_size), selected_color_rgb)
            output_mosaic_img.paste(color_tile, (x, y))

            # 指示書用に色名を記録
            mosaic_color_names_grid[tile_row_idx][tile_col_idx] = target_color_names[best_color_index]
            tile_col_idx += 1
        tile_row_idx += 1

    # 保存ディレクトリ準備
    output_dir = "mosaic_sheets"
    if os.path.exists(output_dir):
        shutil.rmtree(output_dir)
    os.makedirs(output_dir, exist_ok=True)

    # --- 用紙サイズ・向きに応じたExcelセル寸法（列幅・行高）の自動計算 ---
    selected_paper = PAPER_SIZES_MM.get(paper_size_name, PAPER_SIZES_MM["A4"])
    
    if paper_orientation == "ヨコ":
        printable_width_mm = selected_paper["height"]
        printable_height_mm = selected_paper["width"]
    else: # タテ
        printable_width_mm = selected_paper["width"]
        printable_height_mm = selected_paper["height"]

    # 1マスあたりの目標寸法 (mm) 計算（行番号・列番号のヘッダー枠を考慮）
    tile_width_mm = printable_width_mm / (tiles_per_sheet_horizontal + 1.2)
    tile_height_mm = printable_height_mm / (tiles_per_sheet_vertical + 1.2)

    # mm -> Excel単位に換算
    excel_col_width = max(tile_width_mm / 1.8, 3.0)
    excel_row_height = max(tile_height_mm * 2.83, 12.0)

    # 枠線スタイル設定
    thin_border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )

    sheet_width_px = tiles_per_sheet_horizontal * tile_px_size
    sheet_height_px = tiles_per_sheet_vertical * tile_px_size

    individual_sheet_paths = []
    individual_excel_paths = []

    total_mosaic_tiles_vertical = mosaic_height_px // tile_px_size
    total_mosaic_tiles_horizontal = mosaic_width_px // tile_px_size

    # シートごとの生成ループ
    for row_sheet_idx in range(total_sheets_vertical):
        for col_sheet_idx in range(total_sheets_horizontal):
            # 1. 画像シートの切り出し＆保存
            left_px = col_sheet_idx * sheet_width_px
            top_px = row_sheet_idx * sheet_height_px
            right_px = left_px + sheet_width_px
            bottom_px = top_px + sheet_height_px

            sheet_img = output_mosaic_img.crop((left_px, top_px, right_px, bottom_px))
            sheet_filename = os.path.join(output_dir, f"mosaic_sheet_row{row_sheet_idx+1}_col{col_sheet_idx+1}.png")
            sheet_img.save(sheet_filename)
            individual_sheet_paths.append(sheet_filename)

            # 2. Excel指示書の作成
            excel_filename = os.path.join(output_dir, f"mosaic_sheet_row{row_sheet_idx+1}_col{col_sheet_idx+1}_colors.xlsx")
            individual_excel_paths.append(excel_filename)

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = f"指示書_行{row_sheet_idx+1}_列{col_sheet_idx+1}"

            # ページ印刷設定（用紙サイズIDと向き、1ページ自動適合）
            ws.page_setup.paperSize = selected_paper["excel_code"]
            if paper_orientation == "ヨコ":
                ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
            else:
                ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
            
            ws.sheet_properties.pageSetUpPr.fitToPage = True
            ws.page_setup.fitToWidth = 1
            ws.page_setup.fitToHeight = 1

            start_tile_row = row_sheet_idx * tiles_per_sheet_vertical
            end_tile_row = min(start_tile_row + tiles_per_sheet_vertical, total_mosaic_tiles_vertical)
            start_tile_col = col_sheet_idx * tiles_per_sheet_horizontal
            end_tile_col = min(start_tile_col + tiles_per_sheet_horizontal, total_mosaic_tiles_horizontal)

            # ヘッダー書き込み
            header = ["行/列"] + [f"列{j+1}" for j in range(end_tile_col - start_tile_col)]
            ws.append(header)

            # データ書き込み
            for r in range(start_tile_row, end_tile_row):
                row_data = [f"行{r - start_tile_row + 1}"]
                for c in range(start_tile_col, end_tile_col):
                    row_data.append(mosaic_color_names_grid[r][c])
                ws.append(row_data)

            # 行高とセル書式
            for row in ws.iter_rows():
                ws.row_dimensions[row[0].row].height = excel_row_height
                for cell in row:
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    cell.border = thin_border
                    cell.font = Font(size=9)

            # 列幅
            for col in ws.columns:
                col_letter = openpyxl.utils.get_column_letter(col[0].column)
                ws.column_dimensions[col_letter].width = excel_col_width

            wb.save(excel_filename)

    # 3. ZIPファイル作成
    images_zip_path = os.path.join(output_dir, "mosaic_sheets_images.zip")
    with zipfile.ZipFile(images_zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
        for file_path in individual_sheet_paths:
            zipf.write(file_path, os.path.basename(file_path))

    excels_zip_path = os.path.join(output_dir, "mosaic_sheets_excels.zip")
    with zipfile.ZipFile(excels_zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
        for file_path in individual_excel_paths:
            zipf.write(file_path, os.path.basename(file_path))

    return output_mosaic_img, individual_sheet_paths, individual_excel_paths, images_zip_path, excels_zip_path


# --- Streamlit UI構築 ---

st.title("🧩 カスタマイズ可能モザイクアートクリエーター")
st.markdown("""
元画像をアップロードし、使用する色・シート分割数・タイル数・用紙サイズを設定してモザイクアートと指示書を出力します。  
生成された全画像と全Excel指示書は、それぞれZIPファイルとして一括ダウンロードできます。
""")

col_input, col_config = st.columns([1, 1])

with col_input:
    st.subheader("1. 画像と色の設定")
    uploaded_file = st.file_uploader("ベース画像をアップロード", type=["png", "jpg", "jpeg"])

    num_colors = st.slider("使用する色の数", min_value=2, max_value=20, value=5)

    target_colors_data = []
    st.markdown("**各色のカラーピッカーと名前**")
    
    default_hexes = ["#FF0000", "#00FF00", "#0000FF", "#FFFF00", "#FF00FF", "#00FFFF", "#000000", "#FFFFFF"]
    default_names = ["赤", "緑", "青", "黄", "マゼンタ", "シアン", "黒", "白"]

    for i in range(num_colors):
        c1, c2 = st.columns([1, 2])
        hex_val = default_hexes[i % len(default_hexes)]
        name_val = default_names[i % len(default_names)] if i < len(default_names) else f"色{i+1}"
        
        with c1:
            color = st.color_picker(f"色 {i+1}", hex_val, key=f"color_{i}")
        with c2:
            color_name = st.text_input(f"色 {i+1} の名前", name_val, key=f"name_{i}")
        
        target_colors_data.append((color, color_name))

with col_config:
    st.subheader("2. モザイクアートと用紙の設定")
    
    # 用紙サイズと向きの設定
    c_paper1, c_paper2 = st.columns(2)
    with c_paper1:
        paper_size_name = st.selectbox("印刷する用紙サイズ", ["A4", "A3", "B5", "B4"], index=0)
    with c_paper2:
        paper_orientation = st.selectbox("用紙の向き", ["タテ", "ヨコ"], index=0)

    st.markdown("---")
    total_sheets_vertical = st.slider("モザイクアート全体の縦のシート枚数", 1, 20, 8)
    total_sheets_horizontal = st.slider("モザイクアート全体の横のシート枚数", 1, 20, 12)
    tiles_per_sheet_vertical = st.slider("1枚のシートあたりのタイルの縦の数", 5, 50, 10)
    tiles_per_sheet_horizontal = st.slider("1枚のシートあたりのタイルの横の数", 5, 50, 10)

st.markdown("---")

# 生成ボタン
if st.button("🚀 モザイクアートを生成", type="primary", use_container_width=True):
    if uploaded_file is None:
        st.error("ベース画像をアップロードしてください。")
    elif any(not name.strip() for _, name in target_colors_data):
        st.warning("すべての色の名前を入力してください。")
    else:
        with st.spinner("モザイクアートおよび印刷最適化指示書を生成中..."):
            try:
                base_image = Image.open(uploaded_file)
                
                full_mosaic, img_paths, excel_paths, img_zip, excel_zip = create_mosaic_sheets_from_colors(
                    base_image=base_image,
                    target_colors_data=target_colors_data,
                    total_sheets_vertical=total_sheets_vertical,
                    total_sheets_horizontal=total_sheets_horizontal,
                    tiles_per_sheet_vertical=tiles_per_sheet_vertical,
                    tiles_per_sheet_horizontal=tiles_per_sheet_horizontal,
                    paper_size_name=paper_size_name,
                    paper_orientation=paper_orientation
                )

                st.session_state["full_mosaic"] = full_mosaic
                st.session_state["img_zip"] = img_zip
                st.session_state["excel_zip"] = excel_zip
                st.session_state["excel_paths"] = excel_paths
                st.success("モザイクアートが生成されました！")

            except Exception as e:
                st.error(f"エラーが発生しました: {e}")

# 結果表示エリア
if "full_mosaic" in st.session_state:
    st.subheader("🖼️ 生成結果")
    st.image(st.session_state["full_mosaic"], caption="モザイクアート全体像", use_column_width=True)

    col_dl1, col_dl2 = st.columns(2)
    
    with col_dl1:
        with open(st.session_state["img_zip"], "rb") as f:
            st.download_button(
                label="📦 全てのシート画像をZIPでダウンロード",
                data=f,
                file_name="mosaic_sheets_images.zip",
                mime="application/zip",
                use_container_width=True
            )

    with col_dl2:
        with open(st.session_state["excel_zip"], "rb") as f:
            st.download_button(
                label="📄 全てのExcel色指示書をZIPでダウンロード",
                data=f,
                file_name="mosaic_sheets_excels.zip",
                mime="application/zip",
                use_container_width=True
            )

    st.markdown("---")
    st.subheader("📊 各シートの色指示表プレビュー")
    
    excel_paths = st.session_state["excel_paths"]
    excel_names = [os.path.basename(p) for p in excel_paths]
    selected_excel_name = st.selectbox("表示するシートの指示書を選択", excel_names)

    if selected_excel_name:
        selected_path = next(p for p in excel_paths if os.path.basename(p) == selected_excel_name)
        df = pd.read_excel(selected_path)
        st.dataframe(df, use_container_width=True)
